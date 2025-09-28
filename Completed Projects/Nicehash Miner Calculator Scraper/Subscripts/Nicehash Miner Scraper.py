# WARNING: THIS PROJECT WAS ASSISTED WITH CHATGPT, BECAUSE I CAN'T BE ARSED DEALING WITH SELENIUMS NONSENSE



# Selenium imports
import requests
import xml.etree.ElementTree as ET
import pyautogui

print("Starting Selenium Web Handler")
from selenium import webdriver
from selenium.webdriver.firefox.service import Service as FirefoxService
from selenium.webdriver.firefox.options import Options as FirefoxOptions
from webdriver_manager.firefox import GeckoDriverManager
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, WebDriverException
print("Selenium active!")

# --- Selenium setup (headless, sane defaults) ---
opts = FirefoxOptions()
opts.add_argument("--headless")   # uncomment if you want headless mode
opts.add_argument("--width=1920")
opts.add_argument("--height=1080")
#opts.add_argument("--start-fullscreen")

#driver = webdriver.Firefox(service=FirefoxService(GeckoDriverManager().install()),options=opts)
driver = webdriver.Firefox(service=FirefoxService(), options=opts)

#service = FirefoxService(executable_path=r"C:\path\to\geckodriver.exe")
#driver = webdriver.Firefox(service=service, options=opts)












# Openpyxl imports
import openpyxl
from openpyxl import Workbook, load_workbook
from tempfile import NamedTemporaryFile

FILE_PATH = "data.xlsx"
SHEET_NAME = "Data"
HEADER = ["Device Name", "Daily Raw Income (AUD)", "Daily Electrical Costs (AUD)", "Daily Total Income (AUD)", "URL"]






# Other imports
import time, random, csv, os, re








# My program
filtered_urls = []
wait = WebDriverWait(driver, 20)  # adjust if your pages are slow

def GetMinerURLs():
    url = "https://www.nicehash.com/sitemap.xml"
    response = requests.get(url)
    root = ET.fromstring(response.text)

    # XML namespace for sitemaps
    ns = {"ns": "http://www.sitemaps.org/schemas/sitemap/0.9"}

    # Extract all <loc> elements
    urls = [loc.text for loc in root.findall(".//ns:loc", ns)]

    prefix = "https://www.nicehash.com/profitability-calculator/"
    filtered_urls = [u for u in urls if u.startswith(prefix)]

    if not os.path.exists(FILE_PATH):
        return filtered_urls

    # Open workbook and read row 4
    wb = openpyxl.load_workbook(FILE_PATH)
    sheet = wb.active
    row_num = 4

    # Collect all URLs from row 4, skipping A4
    excel_urls = [
        cell.value
        for cell in sheet['E'][1:]  # skip E1 (header), index 1 == row 2
        if cell.value is not None
    ]

    # Now filter out anything that’s already in Excel
    filtered_urls = [url for url in filtered_urls if url not in excel_urls]

    return filtered_urls
    

def LoadNewSession(URL):
    for attempt in range(3):
        try:
            driver.get(URL)

            # Wait for the cookie banner, and close it
            cookie_btn = wait.until(EC.element_to_be_clickable((By.XPATH, "/html/body/div[1]/div/div[7]/div/div/div[2]/button")))
            cookie_btn.click()

            return
        except Exception:
            pass
    print(f"An error occured while loading new session: {URL}")

def wait_for_delayed_autoscroll(timeout=15, poll=0.1):
    """
    Waits until the page does its delayed auto-scroll, then settles.
    Returns True if detected + stable, False if timed out.
    """
    end = time.time() + timeout
    start_y = driver.execute_script("return window.scrollY;")
    last_y = start_y
    seen_scroll = False
    stable_count = 0

    while time.time() < end:
        y = driver.execute_script("return window.scrollY;")

        if y != last_y:
            seen_scroll = True
            stable_count = 0  # reset stability counter

        if seen_scroll:
            if y == last_y:
                stable_count += 1
                if stable_count >= 3:  # ~0.3s stable if poll=0.1
                    return True

        last_y = y
        time.sleep(poll)

    return False

def _top_fixed_header_height():
    """Best-effort: detect tallest fixed header at top (height in px)."""
    return driver.execute_script("""
        let maxH = 0;
        for (const el of document.querySelectorAll('*')) {
            const s = getComputedStyle(el);
            if (s.position === 'fixed') {
                const top = parseInt(s.top || '0', 10);
                if (top === 0) {
                    const h = el.offsetHeight || 0;
                    if (h > maxH) maxH = h;
                }
            }
        }
        return maxH;
    """)

def smart_scroll_to_center(element, *, margin_px=8, max_steps=20, damping=0.6, extra_offset_px=0):
    """
    Scrolls the window so that `element` is centered vertically.
    - margin_px: how close to center is 'good enough'
    - max_steps: safety cap
    - damping: 0..1; smaller = smaller scroll increments (smoother)
    - extra_offset_px: add if you have known sticky bars etc.
    Returns True if centered (within margin), else False.
    """
    # auto-detect fixed header height, add any extra offset you want
    top_offset = 0
    try:
        top_offset = _top_fixed_header_height()
    except Exception:
        pass
    top_offset += int(extra_offset_px)

    for _ in range(max_steps):
        delta = driver.execute_script("""
            const el = arguments[0];
            const offset = arguments[1] || 0;
            const r = el.getBoundingClientRect();
            const elemCenter = r.top + r.height / 2;
            const viewportCenter = (window.innerHeight / 2) + offset;
            return elemCenter - viewportCenter;   // +ve means element is below center
        """, element, top_offset)

        # Already centered enough?
        if delta is None:
            return False
        if abs(delta) <= margin_px:
            return True

        # Scroll proportionally toward center (direction-aware)
        driver.execute_script("window.scrollBy(0, arguments[0]);", float(delta) * float(damping))
        time.sleep(0.5)  # let layout settle a hair

    return False

def SetValues(Currency, Electrical_Cost):
    # Set the currency type
    currency_button_XPATH = "/html/body/div[1]/div/div[6]/div/div[2]/section[1]/div/div[1]/div[2]/div[1]/div[1]/div[1]/div[1]/div/div[2]/div[1]"
    currency_btn = wait.until(EC.element_to_be_clickable((By.XPATH, currency_button_XPATH)))
    if not wait_for_delayed_autoscroll(timeout=30):
        print("Timeout: page kept moving")
    smart_scroll_to_center(currency_btn)
    currency_btn.click()

    input_box = wait.until(EC.presence_of_element_located((By.XPATH, "//*[@id=\"app\"]/div[6]/div/div[2]/section[1]/div/div[1]/div[2]/div[1]/div[1]/div[1]/div[1]/div/div[2]/div[2]/div[1]/input")))
    input_box.clear()
    input_box.send_keys("AUD")
    input_box.send_keys(Keys.ARROW_DOWN)
    input_box.send_keys(Keys.ENTER)

    # Set the electricity costs
    electricity_cost_btn = wait.until(EC.element_to_be_clickable((By.XPATH, "/html/body/div[1]/div/div[6]/div/div[2]/section[1]/div/div[1]/div[2]/div[1]/div[1]/div[1]/div[1]/div/div[2]/div[1]")))
    electricity_cost_btn.click()

    input_box = wait.until(EC.presence_of_element_located((By.XPATH, "//*[@id=\"app\"]/div[6]/div/div[2]/section[1]/div/div[1]/div[2]/div[1]/div[1]/div[1]/div[2]/div/div[2]/input")))
    input_box.clear()
    input_box.send_keys(str(Electrical_Cost))

def CalculateProfit():
    calculate_btn = wait.until(EC.element_to_be_clickable((By.XPATH, "//*[@id=\"app\"]/div[6]/div/div[2]/section[1]/div/div[1]/div[2]/div[1]/div[1]/div[4]/button")))
    calculate_btn.click()

def _atomic_save(wb, path):
    """Write safely so a crash won’t corrupt the file."""
    with NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        tmp_name = tmp.name
        wb.save(tmp_name)
    os.replace(tmp_name, path)  # atomic on most OSes


def get_sheet():
    """Open file if it exists; else create with header once."""
    if os.path.exists(FILE_PATH):
        wb = load_workbook(FILE_PATH)
        ws = wb[SHEET_NAME] if SHEET_NAME in wb.sheetnames else wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = SHEET_NAME
        ws.append(HEADER)     # write header in first row
        _atomic_save(wb, FILE_PATH)

    return wb, ws


def append_row(values):
    """
    Append a new row and save immediately.
    `values` is a list matching HEADER order, e.g. ["foo", 123, 4.56, "ok"]
    """
    wb, ws = get_sheet()
    ws.append(values)
    _atomic_save(wb, FILE_PATH)


def write_cell(row_idx, col_idx, value):
    """
    Update a specific cell and save immediately.
    row_idx: 1-based (1 is header row)
    col_idx: 1-based (1 is column A)
    """
    wb, ws = get_sheet()
    ws.cell(row=row_idx, column=col_idx, value=value)
    _atomic_save(wb, FILE_PATH)

def cleanse_numeric(s: str) -> str:
    """Keep only digits, '.', and '-' from a string."""
    return re.sub(r'[^0-9\.\-]', '', s)

def cell_text(row, col, table_index=1):
    # If there can be multiple such tables, use table_index to pick the Nth table.
    table = f"(//div[contains(@class,'mb16')]//table[" \
            f"contains(concat(' ', normalize-space(@class), ' '), ' nh-new ') and " \
            f"contains(concat(' ', normalize-space(@class), ' '), ' desktop ')])[{table_index}]"
    xpath = f"{table}//tbody/tr[{row}]/td[{col}]"
    el = WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH, xpath)))
    return el.text.strip()

def CollectAndSaveResultingData(URL):
    INCOME_XPATH = "/html/body/div[1]/div/div[6]/div[2]/div[2]/section[1]/div/div[1]/div[2]/div[2]/div/div[3]/table/tbody/tr[1]/td[2]"
    ELECTRICITY_COSTS_XPATH = "/html/body/div[1]/div/div[6]/div[2]/div[2]/section[1]/div/div[1]/div[2]/div[2]/div/div[3]/table/tbody/tr[2]/td[2]"
    TOTAL_XPATH = "/html/body/div[1]/div/div[6]/div[2]/div[2]/section[1]/div/div[1]/div[2]/div[2]/div/div[3]/table/tbody/tr[3]/td[2]"
    DEVICE_XPATH = "//*[@id=\"app\"]/div[6]/div/div[2]/section[1]/div/div[1]/div[2]/div[1]/div[1]/div[2]/div[1]/div/div[2]/div[1]/span"
    
    #incomeValues = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, INCOME_XPATH))).text.split("\n")
    #electricalCostsValues = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, ELECTRICITY_COSTS_XPATH))).text.split("\n")
    #totalIncomeValues = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, TOTAL_XPATH))).text.split("\n")
    deviceName = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, DEVICE_XPATH))).text
    
    incomeValues = cell_text(1,2).split("\n")
    electricalCostsValues = cell_text(2,2).split("\n")
    totalIncomeValues = cell_text(3,2).split("\n")
    
    append_row([
        deviceName,
        cleanse_numeric(incomeValues[1]),
        cleanse_numeric(electricalCostsValues[1]),
        cleanse_numeric(totalIncomeValues[1]),
        URL
    ])
    
# Run the script
URLs = GetMinerURLs()
counter = 0
for i in URLs:
    print(str(counter) + "/" + str(len(URLs)) + ": " + i)
    counter += 1
    
    LoadNewSession(i)
    SetValues("AUD",0.23)
    CalculateProfit()
    wait_for_delayed_autoscroll()
    CollectAndSaveResultingData(i)

    driver.delete_all_cookies()
    driver.execute_script("window.localStorage.clear(); window.sessionStorage.clear();")
driver.quit()

print("No URLs Left, Process Completed!")




































