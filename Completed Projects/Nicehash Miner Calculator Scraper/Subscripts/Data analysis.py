import openpyxl,os,time
from openpyxl.utils import get_column_letter

FILE_PATH = "data.xlsx"
SHEET_NAME = "Data"

if not os.path.exists(FILE_PATH):
    raise Exception("No input file \"data.xlsx\" found!")
print("Found input file")

# Open workbook and read row 4
wb = openpyxl.load_workbook(FILE_PATH)
sheet = wb.active

skipFirst = True
allDataPoints = []
for cell in sheet:
    dataPoint = [
        cell[0].value,
        cell[1].value,
        cell[2].value,
        cell[3].value
    ]

    if skipFirst:
        skipFirst = False
        continue
    allDataPoints.append(dataPoint)

def printArray(arr):
    for i in arr:
        print(i)

def hdr():
    return ["Device Name", "Daily Raw Income (AUD)", "Daily Electrical Costs (AUD)", "Daily Total Income (AUD)"]

CPU_Data         = [hdr()]
NVIDIA_Data      = [hdr()]
INTEL_Data       = [hdr()]
AMD_Data         = [hdr()]
BITMAIN_Data     = [hdr()]
MICROBT_Data     = [hdr()]
INNOSILICON_Data = [hdr()]
GOLDSHELL_Data   = [hdr()]
ICERIVER_Data    = [hdr()]
CANAAN_Data      = [hdr()]
Remaining_Data   = [hdr()]
for r in allDataPoints:
    name = (r[0] or "").upper()
    
    if   "CPU"         in name: CPU_Data.append(r)
    elif "NVIDIA"      in name: NVIDIA_Data.append(r)
    elif "INTEL"       in name: INTEL_Data.append(r)
    elif "AMD"         in name: AMD_Data.append(r)
    elif "BITMAIN"     in name: BITMAIN_Data.append(r)
    elif "MICROBT"     in name: MICROBT_Data.append(r)
    elif "INNOSILICON" in name: INNOSILICON_Data.append(r)
    elif "GOLDSHELL"   in name: GOLDSHELL_Data.append(r)
    elif "ICERIVER"    in name: ICERIVER_Data.append(r)
    elif "CANAAN"      in name: CANAAN_Data.append(r) 
    else:                       Remaining_Data.append(r)

def auto_fit_columns(ws):
    for col in ws.columns:
        max_length = 0
        column = col[0].column  # Get the column index (1-based)
        col_letter = get_column_letter(column)
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = (max_length + 2)  # a little padding
        ws.column_dimensions[col_letter].width = adjusted_width
        
def write_table(wb, title, table):
    ws = wb.create_sheet(title=title)
    for row in table:
        ws.append(row)
    auto_fit_columns(ws)
    
def append_rows(ws, rows):
    """rows must be an iterable of row-lists; for a single row, wrap as [row]."""
    for r in rows:
        ws.append(r)

tables = {
    "CPU": CPU_Data,
    "NVIDIA": NVIDIA_Data,
    "INTEL": INTEL_Data,
    "AMD": AMD_Data,
    "BITMAIN": BITMAIN_Data,
    "MICROBT": MICROBT_Data,
    "INNOSILICON": INNOSILICON_Data,
    "GOLDSHELL": GOLDSHELL_Data,
    "ICERIVER": ICERIVER_Data,
    "CANAAN": CANAAN_Data,
    "Remaining": Remaining_Data,
    "Sorted - Total Profit": [hdr()],
    "Sorted - Raw Daily Income": [hdr()]
}

# Purge old data
for sheetname in wb.sheetnames:
    if sheetname != SHEET_NAME:
        del wb[sheetname]
wb.save(FILE_PATH)
wb = openpyxl.load_workbook(FILE_PATH)

# Now load up the data
for title, table in tables.items():
    write_table(wb, title, table)
wb.save(FILE_PATH)

# Now we want to get the data in terms of profitabilities
totalProfitSorted = sorted(allDataPoints, key=lambda x: float(x[3] or 0), reverse=True)
dailyIncomeSorted = sorted(allDataPoints, key=lambda x: float(x[1] or 0), reverse=True)

ws = wb["Sorted - Total Profit"]
for i in totalProfitSorted:
    if float(i[3]) > 0:
        ws.append(i)
auto_fit_columns(ws)

ws = wb["Sorted - Raw Daily Income"]
for i in dailyIncomeSorted:
    ws.append(i)
auto_fit_columns(ws)

wb.save(FILE_PATH)
print("Sheets written and workbook saved")































