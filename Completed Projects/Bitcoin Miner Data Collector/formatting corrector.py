from openpyxl import Workbook,load_workbook
import openpyxl
import time
wb = load_workbook(filename = "incomeValues.xlsx")
ws = wb.active

looper = 0
while(True):
    looper += 1
    index = "A" + str(looper)
    if (ws[index].value == None):
        break

for i in range(2,looper):
    ws["B" + str(i)].value = float(ws["B" + str(i)].value[:-4])
    ws["C" + str(i)].value = float(ws["C" + str(i)].value[:-4])
    ws["D" + str(i)].value = float(ws["D" + str(i)].value[:-4])
    ws["E" + str(i)].value = float(ws["E" + str(i)].value[:-4])
    ws["F" + str(i)].value = float(ws["F" + str(i)].value[:-4])
    ws["G" + str(i)].value = float(ws["G" + str(i)].value[:-4])
    ws["H" + str(i)].value = float(ws["H" + str(i)].value[:-4])
    ws["I" + str(i)].value = float(ws["I" + str(i)].value[:-4])
    ws["J" + str(i)].value = float(ws["J" + str(i)].value[:-4])

wb.save('Data - Complete Formatted Data.xlsx')

def runtimeConditionalFormatting(worksheet, cellLocation):
    print(cellLocation)
    if (worksheet[cellLocation].value > 0):
        worksheet[cellLocation].fill = openpyxl.styles.PatternFill(fgColor="32C832", fill_type = "solid")
    if (worksheet[cellLocation].value == 0):
        worksheet[cellLocation].fill = openpyxl.styles.PatternFill(fgColor="C89600", fill_type = "solid")
    if (worksheet[cellLocation].value < 0):
        worksheet[cellLocation].fill = openpyxl.styles.PatternFill(fgColor="C80000", fill_type = "solid")

def createFilteredFile(contain, AllContainRequired, notContain, AllNotContainRequired, filename):
    filename += ".xlsx"
    
    tempFile = Workbook()
    ws2 = tempFile.active
    
    startingData = []
    startingData.append(ws['A1'].value)
    startingData.append(ws['B1'].value)
    startingData.append(ws['C1'].value)
    startingData.append(ws['D1'].value)
    startingData.append(ws['E1'].value)
    startingData.append(ws['F1'].value)
    startingData.append(ws['G1'].value)
    startingData.append(ws['H1'].value)
    startingData.append(ws['I1'].value)
    startingData.append(ws['J1'].value)
    ws2.append(startingData)
    
    for i in range(2,looper):
        #Iterating over old data for filter
        test1 = False
        test2 = True
        
        for x in contain:
            if (AllContainRequired):
                test1 = test1 and (x in ws["A" + str(i)].value)
            else:
                test1 = test1 or (x in ws["A" + str(i)].value)

        for x in notContain:
            if (AllNotContainRequired):
                test2 = test2 and (x not in ws["A" + str(i)].value)
            else:
                test2 = test2 or (x not in ws["A" + str(i)].value)

        if (len(contain) == 0):
            test1 = True
        if (len(notContain) == 0):
            test2 = True
        
        if (test1 and test2):
            tempData = []
            tempData.append(ws["A" + str(i)].value)
            tempData.append(ws["B" + str(i)].value)
            tempData.append(ws["C" + str(i)].value)
            tempData.append(ws["D" + str(i)].value)
            tempData.append(ws["E" + str(i)].value)
            tempData.append(ws["F" + str(i)].value)
            tempData.append(ws["G" + str(i)].value)
            tempData.append(ws["H" + str(i)].value)
            tempData.append(ws["I" + str(i)].value)
            tempData.append(ws["J" + str(i)].value)
            ws2.append(tempData)

    tempFile.save(filename)
    time.sleep(1)
    
    for y in range(2,ws2.max_row):
        for x in range(1,9):
            if (x == 1):
                runtimeConditionalFormatting(ws2, "B" + str(y))
            if (x == 2):
                runtimeConditionalFormatting(ws2, "C" + str(y))
            if (x == 3):
                runtimeConditionalFormatting(ws2, "D" + str(y))
            if (x == 4):
                runtimeConditionalFormatting(ws2, "E" + str(y))
            if (x == 5):
                runtimeConditionalFormatting(ws2, "F" + str(y))
            if (x == 6):
                runtimeConditionalFormatting(ws2, "G" + str(y))
            if (x == 7):
                runtimeConditionalFormatting(ws2, "H" + str(y))
            if (x == 8):
                runtimeConditionalFormatting(ws2, "I" + str(y))
            if (x == 9):
                runtimeConditionalFormatting(ws2, "J" + str(y))
    tempFile.save(filename)
                
createFilteredFile(["CPU"], False,[], True, "Data - CPUs")
createFilteredFile(["NVIDIA","AMD"], False, ["CPU"], True, "Data - Graphics Cards")
createFilteredFile([], False, ["CPU","NVIDIA","AMD"], True, "Data - ASICs")

























