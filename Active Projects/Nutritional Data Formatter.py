from openpyxl import Workbook,load_workbook
import os

previousNutData = ""
workbook = Workbook()
sheet = workbook.active

if (not os.path.exists('C:/Users/aj200/Desktop/Nutrition Output.xlsx')):
    print("Could not find excel output file. Creating file")
    sheet["A1"] = "Ingredient"
    sheet["B1"] = "Calories per 100g"
    sheet["C1"] = "Calories percentage"
    sheet["D1"] = "Total Fat per 100g"
    sheet["E1"] = "Total Fat percentage"
    sheet["F1"] = "Saturated Fat per 100g"
    sheet["G1"] = "Saturated Fat percentage"
    sheet["H1"] = "Cholesterol per 100g"
    sheet["I1"] = "Cholesterol percentage"
    sheet["J1"] = "Sodium per 100g"
    sheet["K1"] = "Sodium percentage"
    sheet["L1"] = "Potassium per 100g"
    sheet["M1"] = "Potassium percentage"
    sheet["N1"] = "Total Carbohydrate per 100g"
    sheet["O1"] = "Total Carbohydrate percentage"
    sheet["P1"] = "Dietary Fiber per 100g"
    sheet["Q1"] = "Dietary Fiber percentage"
    sheet["R1"] = "Sugar per 100g"
    sheet["S1"] = "Sugar percentage"
    sheet["T1"] = "Protein per 100g"
    sheet["U1"] = "Protein percentage"
    sheet["V1"] = "Vitamin C per 100g"
    sheet["W1"] = "Vitamin C percentage"
    sheet["X1"] = "Calcium per 100g"
    sheet["Y1"] = "Calcium percentage"
    sheet["Z1"] = "Iron per 100g"
    sheet["AA1"] = "Iron percentage"
    sheet["AB1"] = "Vitamin D per 100g"
    sheet["AC1"] = "Vitamin D percentage"
    sheet["AD1"] = "Vitamin B6 per 100g"
    sheet["AE1"] = "Vitamin B6 percentage"
    sheet["AF1"] = "Cobalamin per 100g"
    sheet["AG1"] = "Cobalamin percentage"
    sheet["AH1"] = "Magnesium per 100g"
    sheet["AI1"] = "Magnesium percentage"
    workbook.save(filename='C:/Users/aj200/Desktop/Nutrition Output.xlsx')
    print("File created!")
else:
    print("Found excel output file. Using and appending to file")
    workbook = load_workbook(filename='C:/Users/aj200/Desktop/Nutrition Output.xlsx')
    sheet = workbook.active

def isValidResponse(x):
    forbiddenResponses = [
        "Amount Per %",
        "Nutrition facts %",
        "100 grams %",
        "Daily Value* %",
        " %",
        "%"
    ]
    
    for i in forbiddenResponses:
        if (i == x):
            return False
    return True

print("Scanning excel file for starting index")
running = True
lineIndex = 1
while(running):
    if (not sheet["A" + str(lineIndex)].value):
        print("Starting index found at: A" + str(lineIndex))
        break
    lineIndex += 1

def formatLine(x):
    i = 0
    while(i < len(x)):
        if ((x[i] == "%") and (x[i-1] == " ")):
            x = x[:i-1] + x[i:]
            i -= 1
        if ((x[i] == "g") and (x[i-1] == " ") and (x[i-2].isnumeric())):
            x = x[:i-1] + x[i:]
            i -= 1
        if ((x[i] == "g") and (x[i-2] == " ") and (x[i-3].isnumeric())):
            x = x[:i-2] + x[i-1:]
            i -= 2
        i += 1
    return x

def getSplitFormattedLine(x):
    x = x.replace(" %","")
    x = formatLine(x)
    
    splitData = x.split(" ")
    ingredientName = ""
    ingredientAmount = ""
    ingredientPercentage = ""
    
    for i in splitData:
        isValue = False
        hasB = False #This variable is used because of Vitamin B6 tripping up and thinking it is a value
        for letter in i:
            if (letter.isnumeric()):
                isValue = True
            if (letter == "B"):
                hasB = True
        if (not isValue):
            if (ingredientName != ""):
                ingredientName += " "
            ingredientName += i
        else:
            if ("%" in i):
                ingredientPercentage = i
            else:
                if (hasB):
                    if (ingredientName != ""):
                        ingredientName += " "
                    ingredientName += i
                else:
                    ingredientAmount = i
    return [ingredientName,ingredientAmount,ingredientPercentage]

def getAllRecordedIngredients():
    global sheet

    results = []
    line = 2
    while(sheet.cell(row=line, column=1).value):
        results.append(sheet.cell(row=line, column=1).value.replace("\n",""))
        line += 1
    return results

print()
with open('C:/Users/aj200/Desktop/Nutrition Input.txt','r') as nutFile:
    while(True):
        #Firstly, read from the user file to get the input data
        nutFile.seek(0)
        x = nutFile.readlines()
        while(x == previousNutData):
            nutFile.seek(0)
            x = nutFile.readlines()
        previousNutData = x[:]
        x[3] = x[3].replace("\n","")
        
        #Do a quick check to see if we have already done this particular ingredient
        allIngredients = getAllRecordedIngredients()
        for z in range(0,len(allIngredients)):
            print(allIngredients[z].replace("\n",""))
            if (str(x[3].replace("\n","")) == str(allIngredients[z].replace("\n",""))):
                print("Found pre-existing ingredient by that name. Overwriting old data")
                lineIndex = z + 2
        currentLine = lineIndex
        
        #Now split up the data, get rid of nasty characters, and throw  out known false results
        resultingValues = []
        for i in x:
            z = i.split("%")
            for y in range(0,len(z)):

                z[y] = z[y].replace("\n"," ").replace("\t"," ").replace("  "," ") + "%"
                if (z[y] == " %"):
                    z.pop(y)
            for y in range(0,len(z)):
                if (z[y][0] == " "):
                    z[y] = z[y][1:]
                if (isValidResponse(z[y])):
                    resultingValues.append(z[y])

        #Remove the first 3 values, they are just the name of the food
        resultingValues.pop(0)
        resultingValues.pop(0)
        sources = resultingValues.pop(0)
        resultingValues[0] = resultingValues[0].replace(" %","cal")
        
        #Now that we have "Sort of" formatted string, split them up into usable variables and then place them into the excel spreadsheet
        sheet["A" + str(lineIndex)] = x[3]
        for i in resultingValues:
            tmp = getSplitFormattedLine(i)
            tmp[0] = tmp[0].title()
            print(tmp)
            if(tmp[1] != ""):
                columbIndex = 1
                while(sheet.cell(row=1, column=columbIndex).value):
                    if (str(sheet.cell(row=1, column=columbIndex).value) == str(tmp[0] + " per 100g")):
                        break
                    columbIndex += 1
                sheet.cell(row=lineIndex, column=columbIndex).value = tmp[1]
                sheet.cell(row=1, column=columbIndex).value = tmp[0]
            if(tmp[2] != ""):
                columbIndex = 1
                while(sheet.cell(row=1, column=columbIndex).value):
                    if (str(sheet.cell(row=1, column=columbIndex).value) == str(tmp[0] + " percentage")):
                        break
                    columbIndex += 1
                sheet.cell(row=lineIndex, column=columbIndex).value = tmp[2]
                sheet.cell(row=1, column=columbIndex).value = tmp[0]
            lineIndex = currentLine
        
        #This code automatically resizes the cells for easy viewing. I did not write this (Copy and paste from stack overflow), but it seems to work...
        for col in sheet.columns:
            max_length = 0
            column = col[0].column_letter # Get the column name
            for cell in col:
                try: # Necessary to avoid error on empty cells
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = max_length + 1
            sheet.column_dimensions[column].width = adjusted_width
        
        #Finally save the file to disk, ready for the next set of data to be entered
        workbook.save(filename='C:/Users/aj200/Desktop/Nutrition Output.xlsx')
        lineIndex += 1
        print("Successfully saved ingredient: " + x[3])



























