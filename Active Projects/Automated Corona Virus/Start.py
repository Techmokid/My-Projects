#https://coronavirus.jhu.edu/map.html?fbclid=IwAR3iMsAKkkl_0p3y4ohwSt2ChUFUEdBUu_Dfn7O_kpKA3itNuOsVOx9vduM
#https://hgis.uw.edu/virus/?fbclid=IwAR2UdNKgxeBKRNoAK5G59kuEGkQgo73-dFkDH_r7pO6h1AntQBSd4rUYJaw
#https://www.worldometers.info/coronavirus/?fbclid=IwAR1QqQBkcGY3I2r0f4lHSQibBXZDJWaQtP6yQhPVuHn-fJo1JiedXEDDDVM

import requests,time,os,urllib,shutil,openpyxl
from openpyxl.styles import NamedStyle, Font, Border, Side, PatternFill
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
from datetime import datetime, timedelta

def splitTextToLines(text):
    result = [""]
    pos = 0
    for i in text:
        if (i != '\n'):
            result[pos] += i
        else:
            result.append("")
            pos += 1
    return result

def getValues(valIn,key,endPoint="</div"):
    divs = 0
    running = False
    myString = ""
    results = []
    for i in valIn:
        if(key in i):
            running = True
        if (("<div" in i) and (running)):
            divs += 1
        if ((divs > 0) and (endPoint in i)):
            divs -= 1
            if (divs == 0):
                running = False
                result.append(myString)
                myString = ""
        if (divs > 0):
            myString += i + '\n'
    return results

def getLineOfValue(data,key):
    looper = 0
    for i in data:
        if (key in i):
            return looper
        looper += 1

def getListSubstring(data,start,end):
    looper = 0
    result = []
    for i in data:
        if ((looper > start) and (looper < end)):
            result.append(i)
        looper += 1
    return result

def getListToString(lst,key):
    result = []
    for i in lst:
        if (key in i):
            return result
        result.append(i)
    return None

def download_url(url, save_Path):
    with urllib.request.urlopen(url) as dl_file:
        with open(save_path, 'wb') as out_file:
            out_file.write(dl_file.read())

def HasInternetConnection():
    try:
        urllib.request.urlopen('http://google.com')
        return True
    except:
        return False

while(HasInternetConnection() == False):
    print("Waiting for internet connection...")
    time.sleep(10)

def printList(text):
    for i in text:
        print("----------------------------------------------------------------------")
        time.sleep(1)
        print(i)
        input()

def getValueFromList(data,startPoint,key,endPoint):
    result = []
    currentList = ""
    running = 0
    for line in data:
        #print("Running:",running,"\t\tLine:",line)
        #time.sleep(0.5)
        if (startPoint in line):
            running += 1
        if (running != 0):
            currentList += line + "\n"
        if (endPoint in line):
            running -= 1
            if (running == 0):
                if (key in currentList):
                    result.append(currentList)
                currentList = ""
    return result

def reformatData(rawData):
    worldDataList = []
    
    for countryData in rawData:
        countryDataList = []
        lineType = 0
        for line in splitTextToLines(countryData):
            if ("text-align:right" in line):
                value = line[line.find(">")+1:-5]
                lineType += 1
                
                if ((lineType > 0) and (lineType < 13)):
                    countryDataList.append(value)
                elif (lineType == 13):
                    temp2 = value[value.find(">") + 1:]
                    countryDataList.append(temp2[:temp2.find("<")])
                    
                    temp = value[value.find("\"")+1:]
                    CountryDirectory = temp[:temp.find("\"")]
                    RawCountryName = CountryDirectory[CountryDirectory[:-1].rfind("/") + 1:-1]
                    PureCountryName = RawCountryName.split("-")[0]
                    countryDataList.append(PureCountryName)
        worldDataList.append(countryDataList)
    return worldDataList

#Run data collection services
print("[Data Collection]:   Collecting data from \"World O Meters\"...")
rawSiteData_1 = splitTextToLines(requests.get("https://www.worldometers.info/coronavirus/?fbclid=IwAR1QqQBkcGY3I2r0f4lHSQibBXZDJWaQtP6yQhPVuHn-fJo1JiedXEDDDVM").text)
scrapedSiteData_1 = getLineOfValue(rawSiteData_1,"<table id=\"main_table_countries_today\"")

print("[Data Collection]:   Sorting data by country...")
temp_1 = getListSubstring(rawSiteData_1,scrapedSiteData_1,len(rawSiteData_1) - 1)
dataList = getValueFromList(temp_1,"<tr","style=\"font-weight:","</tr>")

print("[Data Collection]:   Reformatting data...")
worldDataList = reformatData(dataList)

#Create the required database
print("[Database Creation]: Creating required files...")
path = os.getcwd() + "/Retrieved Data/"
if (os.path.exists(path)):
    while(True):
        try:
            shutil.rmtree(path)
            break
        except:
            print("Cannot access required folder. Make sure nothing is using the following folder:")
            print(" - " + path)
            print("\nRetrying in 30 seconds...\n\n")
            time.sleep(30)

os.mkdir(path)

print("[Database Creation]: Saving raw data to file...")

book = openpyxl.load_workbook('C:/Users/aj200/Documents/GitHub/My-Projects/My-Projects/Active Projects/Automated Corona Virus/Active-Coronavirus-cases.xlsx')
sheet = book.active
sheet.title = "Worldwide Active Cases"

activeCaseCount = 0
for countryDataList in worldDataList:
    temp = ""
    for i in countryDataList[6]:
        if (i.isdigit()):
            temp += i
    if (len(temp) != 0):
        activeCaseCount += int(temp)

def cleanDigit(x):
    result = ""
    for i in x:
        if (i.isdigit()):
            result += i
    return result

tempNow = str(int(datetime.today().strftime("%d")))
if (tempNow[-1] == "1"):
    tempNow += "st"
elif (tempNow[-1] == "2"):
    tempNow += "nd"
elif (tempNow[-1] == "3"):
    tempNow += "rd"
else:
    tempNow += "th"
tempNow += " of " + datetime.today().strftime("%B")

activeCaseCount = "#VALUE"
oldDate = sheet["C4"].value.split(':')[1]

date_time_obj = datetime.strptime(cleanDigit(oldDate.split('of')[0][1:]) + oldDate.split('of')[1][:4], '%d %b')
date_time_future = datetime.today() + timedelta(weeks=1)
tempPlusWeek = str(int(date_time_future.strftime("%d")))
if (tempPlusWeek[-1] == "1"):
    tempPlusWeek += "st"
elif (tempPlusWeek[-1] == "2"):
    tempPlusWeek += "nd"
elif (tempPlusWeek[-1] == "3"):
    tempPlusWeek += "rd"
else:
    tempPlusWeek += "th"
tempPlusWeek += " of " + datetime.today().strftime("%B")

sheet["C4"] = "Next update: " + tempPlusWeek
sheet["B7"].value = "=SUM(B9:B" + str(sheet.max_row) + ")"
sheet["C7"].value = "=SUM(C9:C" + str(sheet.max_row) + ")"
sheet["F7"].value = "=SUM(F9:F" + str(sheet.max_row) + ")"
sheet["G7"].value = "=SUM(G9:G" + str(sheet.max_row) + ")"
sheet["H7"].value = "=SUM(H9:H" + str(sheet.max_row) + ")"

for i in range(9,sheet.max_row + 1):
    if (str(sheet["B" + str(i)].value) != "None"):
        sheet["C" + str(i)] = sheet["B" + str(i)].value
        sheet["D" + str(i)] = "=B" + str(i) + "-C" + str(i)
        sheet["E" + str(i)] = "=D" + str(i) + "/C" + str(i)
        sheet["F" + str(i)] = "=SUM(B" + str(i) + ",G" + str(i) + ",H" + str(i) + ")"
        sheet["G" + str(i)] = ""
        sheet["H" + str(i)] = ""
        sheet["I" + str(i)] = "=G" + str(i) + "+H" + str(i)
        sheet["J" + str(i)] = "=I" + str(i) + "/F" + str(i)
        
        sheet["B" + str(i)] = ""
        #Here is where we actually insert the data
        #0 - total cases[F?]
        #1 - new cases
        #2 - total deaths[H?]
        #3 - new deaths
        #4 - total recovered[G?]
        #5 - ?
        #6 - active cases[B?]
        #13 - name [A?]

        #We require B, G, H (And technically for identification "A")




























#Here we are setting up conditional formatting
Reinfected = PatternFill(start_color='E06666', end_color='E06666', fill_type='solid')
Recovered = PatternFill(start_color='6AA84F', end_color='6AA84F', fill_type='solid')
GoodDecrease = PatternFill(start_color='34A853', end_color='34A853', fill_type='solid')
badIncrease = PatternFill(start_color='CC0000', end_color='CC0000', fill_type='solid')

sheet2 = book["Past Worldwide Graphs"]
sheet2["A" + str(sheet2.max_row)].value = tempPlusWeek
#sheet2["B" + str(sheet2.max_row)].value = "='Worldwide Active Cases'!D7"
sheet2["I" + str(sheet2.max_row + 1)].value = tempPlusWeek
#sheet2["J" + str(sheet2.max_row)].value = sheet["B7"].value

sheet2["B" + str(sheet2.max_row - 1)].style = 'Comma'
sheet2["J" + str(sheet2.max_row)].style = 'Comma'

bd = Side(style='thin', color="000000")
sheet2["A" + str(sheet2.max_row - 1)].border = Border(left=bd, top=bd, right=bd, bottom=bd)
sheet2["B" + str(sheet2.max_row - 1)].border = Border(left=bd, top=bd, right=bd, bottom=bd)
sheet2["I" + str(sheet2.max_row)].border = Border(left=bd, top=bd, right=bd, bottom=bd)
sheet2["J" + str(sheet2.max_row)].border = Border(left=bd, top=bd, right=bd, bottom=bd)

#set to font size 10
sheet2["B" + str(sheet2.max_row - 1)].font = Font(size=10)
sheet2["J" + str(sheet2.max_row)].font = Font(size=10)

sheet["A1"].value = "=\"\"\"Was\"\" numbers are from the " + oldDate + ". As of the " + tempNow + " the world has passed \"&B7&\" active cases. \""

book.save(path + "worldometers_data.xlsx")





































